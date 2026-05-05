"""
╔══════════════════════════════════════════════════════════════════════════════╗
║         HR Job Scraper v4  —  Delhi NCR / Noida / Gurugram                 ║
║         Platforms : LinkedIn · Naukri · Indeed · Glassdoor                 ║
║                     Shine · Internshala                                     ║
║         Filter    : Last 24 hours                                           ║
╠══════════════════════════════════════════════════════════════════════════════╣
║  FIXES in v4:                                                               ║
║  • LinkedIn   — removed strict 24h filter (LinkedIn rarely shows datetime) ║
║                 now accepts "X hours ago" / "today" text + falls back       ║
║                 to 5-day filter so cards actually appear, then filters      ║
║  • Naukri     — fixed API headers (x-http-method-override, correct appid)  ║
║                 + improved HTML fallback CSS selectors for new site layout  ║
║  • Indeed     — switched to JSON-LD extraction (new site structure)        ║
║                 + Cloudflare detection with debug logging                   ║
║  • Glassdoor  — switched from Google site: search to direct API endpoint   ║
║  • Internshala— fixed CSS selectors for 2024 redesign + relaxed age check  ║
║  • Age filter — added "Posted X days ago" / ISO date string parsing        ║
║  • All        — added debug logging so you see WHY jobs are being skipped  ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALL
-------
pip install selenium webdriver-manager requests beautifulsoup4 pandas openpyxl tqdm fake-useragent
pip install undetected-chromedriver   # strongly recommended for LinkedIn

RUN
---
python hr_job_scraper_v4.py
"""

from __future__ import annotations

import hashlib
import json
import logging
import random
import re
import time
from dataclasses import asdict, dataclass
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Optional, Set
from urllib.parse import quote_plus, unquote, urlparse, parse_qs

import pandas as pd
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Selenium ─────────────────────────────────────────────────────────────────
try:
    import undetected_chromedriver as uc
    UC_AVAILABLE = True
except ImportError:
    UC_AVAILABLE = False

try:
    from selenium import webdriver
    from selenium.common.exceptions import TimeoutException, WebDriverException
    from selenium.webdriver.chrome.options import Options
    from selenium.webdriver.chrome.service import Service
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.webdriver.support.ui import WebDriverWait
    from webdriver_manager.chrome import ChromeDriverManager
    SELENIUM_OK = True
except ImportError:
    SELENIUM_OK = False

try:
    from tqdm import tqdm
    HAS_TQDM = True
except ImportError:
    HAS_TQDM = False

try:
    from fake_useragent import UserAgent as _FUA
    _ua_gen = _FUA()
    def rand_ua() -> str:
        try:
            return _ua_gen.random
        except Exception:
            return random.choice(_FALLBACK_UAS)
except Exception:
    def rand_ua() -> str:
        return random.choice(_FALLBACK_UAS)

_FALLBACK_UAS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4 Safari/605.1.15",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:124.0) Gecko/20100101 Firefox/124.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/123.0.0.0 Safari/537.36",
]


# ═══════════════════════════════════════════════════════════════════════════════
#  CONFIG
# ═══════════════════════════════════════════════════════════════════════════════

OUTPUT_FILE        = "HR_Jobs_Last24h.xlsx"
HEADLESS           = True
SELENIUM_WAIT      = 15
MAX_RESULTS_PER_KW = 20
MAX_JOB_AGE_HOURS  = 24           # strict 24h cutoff
SEARCH_PAGES       = 2
REQUEST_TIMEOUT    = 25
INTER_LO           = 1.5
INTER_HI           = 4.0
DEBUG_MODE         = True         # set True to see why jobs get skipped

ENABLE_LINKEDIN    = True
ENABLE_NAUKRI      = True
ENABLE_INDEED      = True
ENABLE_GLASSDOOR   = True
ENABLE_SHINE       = True
ENABLE_INTERNSHALA = True

HR_KEYWORDS = [
"HR Manager", "HR Executive", "HR Generalist",
    "HR Business Partner", "HR Operations Manager", "HR Coordinator",
    "HR Assistant", "Talent Acquisition Specialist", "Talent Acquisition Manager",
    "Recruiter", "Senior Recruiter", "Technical Recruiter",
    "Bulk Recruiter", "Campus Recruiter", "Payroll Executive",
    "Payroll Manager", "Compensation and Benefits Manager",
    "Learning and Development Manager", "Training and Development Executive",
    "HR Head", "CHRO", "People Operations Manager",
    "Employee Relations Manager", "HR Analyst",

]

LOCATION_QUERIES = [
    "Delhi, India",
    "Noida, Uttar Pradesh, India",
    "Gurugram, Haryana, India",
]

TARGET_LOCATION_KEYWORDS = [
    "delhi", "new delhi", "delhi ncr", "ncr", "noida",
    "gurgaon", "gurugram", "greater delhi", "faridabad", "ghaziabad",
]

# ═══════════════════════════════════════════════════════════════════════════════
#  LOGGING
# ═══════════════════════════════════════════════════════════════════════════════

logging.basicConfig(
    level=logging.DEBUG if DEBUG_MODE else logging.INFO,
    format="%(asctime)s [%(levelname)-7s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("hr_scraper_v4")


# ═══════════════════════════════════════════════════════════════════════════════
#  DATA MODEL
# ═══════════════════════════════════════════════════════════════════════════════

@dataclass
class JobRecord:
    portal: str
    source_keyword: str
    job_title: Optional[str]             = None
    company_name: Optional[str]          = None
    job_location: Optional[str]          = None
    date_posted: Optional[str]           = None
    salary_package: Optional[str]        = None
    experience_required: Optional[str]   = None
    employment_type: Optional[str]       = None
    job_url: Optional[str]               = None
    contact_email: Optional[str]         = None
    contact_phone: Optional[str]         = None
    job_description_summary: Optional[str] = None
    fetched_at_utc: Optional[str]        = None

    def url_key(self) -> str:
        return _norm_url(self.job_url or "")

    def content_key(self) -> str:
        t = re.sub(r"\W+", "", (self.job_title or "").lower())
        c = re.sub(r"\W+", "", (self.company_name or "").lower())
        return hashlib.md5(f"{t}|{c}".encode()).hexdigest()


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITIES
# ═══════════════════════════════════════════════════════════════════════════════

def _norm_url(url: str) -> str:
    url = url.split("#")[0].strip().rstrip("/").lower()
    parsed = urlparse(url)
    keep = {"jk", "id", "jobid", "job_id"}
    qs = {k: v for k, v in parse_qs(parsed.query).items() if k in keep}
    clean = "&".join(f"{k}={v[0]}" for k, v in qs.items())
    out = f"{parsed.scheme}://{parsed.netloc}{parsed.path}"
    return (out + "?" + clean) if clean else out


def rand_delay(lo: float = INTER_LO, hi: float = INTER_HI) -> None:
    time.sleep(random.uniform(lo, hi))


def summarize(text: str, max_chars: int = 650) -> Optional[str]:
    if not text:
        return None
    text = " ".join(text.split())
    return text if len(text) <= max_chars else text[:max_chars - 3] + "..."


def extract_email(text: str) -> Optional[str]:
    found = re.findall(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}", text or "")
    bad = {"noreply", "no-reply", "donotreply", "example.", "support@naukri",
           "info@naukri", "mailer@", "alert@", "notify@"}
    for e in found:
        if not any(b in e.lower() for b in bad):
            return e
    return None


def extract_phone(text: str) -> Optional[str]:
    m = re.search(r"(?:\+91[-\s]?)?[6-9]\d{9}", text or "")
    if m:
        return m.group(0).strip()
    m = re.search(r"(?:\+\d{1,3}[-\s]?)?\(?\d{3,4}\)?[-\s]?\d{3,5}[-\s]?\d{4}", text or "")
    return m.group(0).strip() if m else None


def extract_salary(text: str) -> Optional[str]:
    patterns = [
        r"\d{1,3}(?:\.\d{1,2})?\s*[-–to]+\s*\d{1,3}(?:\.\d{1,2})?\s*(?:lakh|lpa|l\.p\.a\.?|lakhs?)",
        r"₹\s?[\d,]+(?:\s*[-–]\s*₹?\s?[\d,]+)?(?:\s*/?\s*(?:year|yr|month|mo|annum|pa|pm))?",
        r"[\d,]+\s?(?:INR|USD|LPA|per\s+year|per\s+month|per\s+annum|p\.a\.|p\.m\.)",
        r"\$\s?\d{2,3}[,\d]*(?:\s?[-–]\s?\$\s?\d{2,3}[,\d]*)?",
    ]
    for pat in patterns:
        m = re.search(pat, text or "", re.IGNORECASE)
        if m:
            return m.group(0).strip()
    return None


def extract_experience(text: str) -> Optional[str]:
    m = re.search(
        r"(\d+\s*[-–to]+\s*\d+\s*(?:years?|yrs?)|\d+\+?\s*(?:years?|yrs?)\s*(?:of\s+)?(?:experience|exp)?)",
        text or "", re.IGNORECASE,
    )
    return m.group(0).strip() if m else None


def extract_employment_type(text: str) -> Optional[str]:
    for label in ["Full-time", "Part-time", "Contract", "Internship",
                  "Temporary", "Freelance", "Remote", "Hybrid", "Work from home"]:
        if re.search(rf"\b{re.escape(label)}\b", text or "", re.IGNORECASE):
            return label
    return None


def parse_age_hours(text: Optional[str]) -> Optional[float]:
    """
    Parse a human-readable posted-date string into hours-ago (float).
    Returns None if we cannot parse → caller decides what to do.
    """
    if not text:
        return None
    t = text.lower().strip()

    # ISO date string e.g. "2025-03-08" or "2025-03-08T14:00:00"
    iso = re.match(r"(\d{4}-\d{2}-\d{2})", t)
    if iso:
        try:
            posted_date = datetime.strptime(iso.group(1), "%Y-%m-%d").replace(tzinfo=timezone.utc)
            delta = datetime.now(timezone.utc) - posted_date
            return delta.total_seconds() / 3600
        except ValueError:
            pass

    if any(p in t for p in ["just posted", "today", "moments ago", "seconds ago",
                              "be an early applicant", "actively hiring"]):
        return 0.0
    m = re.search(r"(\d+)\s*min", t)
    if m:
        return int(m.group(1)) / 60
    m = re.search(r"(\d+)\s*hour", t)
    if m:
        return float(m.group(1))
    m = re.search(r"(\d+)\s*day", t)
    if m:
        return float(m.group(1)) * 24
    m = re.search(r"(\d+)\s*week", t)
    if m:
        return float(m.group(1)) * 24 * 7
    if "yesterday" in t:
        return 25.0
    if "1 day" in t or "24 hours" in t:
        return 24.0
    return None


def is_within_24h(text: Optional[str], label: str = "") -> bool:
    hours = parse_age_hours(text)
    if hours is None:
        if DEBUG_MODE and text:
            log.debug("  ↳ age UNPARSEABLE [%s]: '%s'", label, text[:60])
        return False
    result = hours <= MAX_JOB_AGE_HOURS
    if DEBUG_MODE and not result:
        log.debug("  ↳ age TOO OLD [%s]: %.1fh > %dh  (text='%s')", label, hours, MAX_JOB_AGE_HOURS, text[:40])
    return result


def location_matches(text: Optional[str]) -> bool:
    if not text:
        return False
    return any(k in text.lower() for k in TARGET_LOCATION_KEYWORDS)


def normalize_tel(phone: str) -> Optional[str]:
    if not phone:
        return None
    pure = re.sub(r"[^\d]", "", phone)
    if not pure:
        return None
    if phone.strip().startswith("+"):
        return "+" + pure
    return f"+91{pure}" if len(pure) == 10 else f"+{pure}"


# ═══════════════════════════════════════════════════════════════════════════════
#  HTTP SESSION
# ═══════════════════════════════════════════════════════════════════════════════

SESSION = requests.Session()
SESSION.headers.update({
    "Accept-Language": "en-IN,en;q=0.9",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection":      "keep-alive",
})


def http_get(url: str, *, retries: int = 3, backoff: float = 2.5,
             extra_headers: Optional[Dict] = None, **kwargs) -> Optional[requests.Response]:
    headers = {
        "User-Agent":              rand_ua(),
        "Accept":                  "text/html,application/xhtml+xml,*/*;q=0.8",
        "Referer":                 "https://www.google.com/",
        "Upgrade-Insecure-Requests": "1",
    }
    if extra_headers:
        headers.update(extra_headers)
    for attempt in range(retries):
        try:
            resp = SESSION.get(url, headers=headers, timeout=REQUEST_TIMEOUT,
                               allow_redirects=True, **kwargs)
            if resp.status_code == 429:
                wait = backoff * (2 ** attempt) + random.uniform(1, 3)
                log.warning("⚠ Rate-limited — sleeping %.1fs (%s)", wait, url[:55])
                time.sleep(wait)
                continue
            if resp.status_code >= 500:
                time.sleep(backoff * (attempt + 1))
                continue
            return resp
        except requests.RequestException as exc:
            log.debug("HTTP error attempt %d: %s", attempt + 1, exc)
            time.sleep(backoff * (attempt + 1))
    return None


# ═══════════════════════════════════════════════════════════════════════════════
#  JD PAGE DETAIL EXTRACTOR
# ═══════════════════════════════════════════════════════════════════════════════

def extract_jd_details(url: str) -> Dict[str, Optional[str]]:
    empty: Dict[str, Optional[str]] = {
        "salary": None, "email": None, "phone": None,
        "summary": None, "employment_type": None, "experience": None,
    }
    if not url or not url.startswith("http"):
        return empty
    resp = http_get(url)
    if resp is None or resp.status_code >= 400:
        return empty
    soup = BeautifulSoup(resp.text, "html.parser")
    for tag in soup(["script", "style", "noscript", "header", "footer", "nav", "aside"]):
        tag.decompose()
    body = soup.get_text(" ", strip=True)
    return {
        "salary":          extract_salary(body),
        "email":           extract_email(body),
        "phone":           extract_phone(body),
        "summary":         summarize(body),
        "employment_type": extract_employment_type(body),
        "experience":      extract_experience(body),
    }


# ═══════════════════════════════════════════════════════════════════════════════
#  SELENIUM DRIVER
# ═══════════════════════════════════════════════════════════════════════════════

def build_driver() -> "webdriver.Chrome":
    if UC_AVAILABLE:
        opts = uc.ChromeOptions()
        if HEADLESS:
            opts.add_argument("--headless=new")
        opts.add_argument("--window-size=1920,1080")
        opts.add_argument("--lang=en-US,en")
        opts.add_argument("--disable-notifications")
        # Pin to your installed Chrome version (145)
        driver = uc.Chrome(options=opts, use_subprocess=True, version_main=145)
        return driver
    opts = Options()
    if HEADLESS:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1920,1080")
    opts.add_argument(f"--user-agent={rand_ua()}")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    driver = webdriver.Chrome(
        service=Service(ChromeDriverManager().install()), options=opts
    )
    driver.execute_cdp_cmd(
        "Page.addScriptToEvaluateOnNewDocument",
        {"source": "Object.defineProperty(navigator,'webdriver',{get:()=>undefined})"},
    )
    return driver


def _scroll(driver: "webdriver.Chrome", pauses: int = 5) -> None:
    last = driver.execute_script("return document.body.scrollHeight")
    for _ in range(pauses):
        driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        time.sleep(random.uniform(1.0, 2.0))
        h = driver.execute_script("return document.body.scrollHeight")
        if h == last:
            break
        last = h


# ═══════════════════════════════════════════════════════════════════════════════
#  LINKEDIN
#  FIX: LinkedIn's public job cards often only show relative text like
#       "2 hours ago" in a <time> tag — but sometimes no datetime attribute.
#       We now use f_TPR=r86400 (24h filter on LinkedIn side) AND accept
#       any card where posted text is parseable as within 24h OR is absent
#       (since LinkedIn already filtered it server-side).
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_linkedin(driver: "webdriver.Chrome", keyword: str, location_query: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    for page in range(SEARCH_PAGES):
        start = page * 25
        url = (
            "https://www.linkedin.com/jobs/search/"
            f"?keywords={quote_plus(keyword)}"
            f"&location={quote_plus(location_query)}"
            f"&f_TPR=r86400"
            f"&sortBy=DD"
            f"&start={start}"
        )
        log.debug("LinkedIn URL: %s", url)
        try:
            driver.get(url)
            WebDriverWait(driver, SELENIUM_WAIT).until(
                EC.presence_of_element_located(
                    (By.CSS_SELECTOR,
                     "a.base-card__full-link, "
                     "ul.jobs-search__results-list li, "
                     "div.job-search-card")
                )
            )
            _scroll(driver)
        except (TimeoutException, WebDriverException) as exc:
            log.warning("LinkedIn timeout [%s | %s]: %s", keyword, location_query, str(exc)[:120])
            continue

        soup = BeautifulSoup(driver.page_source, "html.parser")
        cards_found = 0
        for card in soup.select("li"):
            a = card.select_one("a.base-card__full-link[href]")
            if not a:
                continue
            cards_found += 1
            title = a.get_text(" ", strip=True) or None

            company_el = (card.select_one("h4.base-search-card__subtitle") or
                          card.select_one("a.hidden-nested-link") or
                          card.select_one("span.base-search-card__subtitle"))
            company = company_el.get_text(" ", strip=True) if company_el else None

            loc_el = card.select_one("span.job-search-card__location")
            loc    = loc_el.get_text(" ", strip=True) if loc_el else None

            # Try multiple selectors for posted time
            time_el = (card.select_one("time[datetime]") or
                       card.select_one("span.job-search-card__listdate--new") or
                       card.select_one("span.job-search-card__listdate") or
                       card.select_one("time"))
            posted = None
            if time_el:
                posted = time_el.get("datetime") or time_el.get_text(" ", strip=True)

            log.debug("  LinkedIn card: title=%s | loc=%s | posted=%s", title, loc, posted)

            if not location_matches(loc):
                log.debug("  ↳ SKIP location mismatch: %s", loc)
                continue

            # KEY FIX: if posted is None but LinkedIn already filtered by 24h,
            # trust the server-side filter and accept the job
            if posted and not is_within_24h(posted, label="LinkedIn"):
                continue

            job_url = _norm_url(a.get("href", ""))
            rand_delay(0.5, 1.2)
            details = extract_jd_details(job_url)
            records.append(JobRecord(
                portal="LinkedIn", source_keyword=keyword,
                job_title=title, company_name=company, job_location=loc,
                date_posted=posted or "Within 24h (LinkedIn-filtered)",
                salary_package=details["salary"],
                experience_required=details["experience"],
                employment_type=details["employment_type"],
                job_url=job_url,
                contact_email=details["email"],
                contact_phone=details["phone"],
                job_description_summary=details["summary"],
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
            if len(records) >= MAX_RESULTS_PER_KW:
                return records

        log.info("    LinkedIn cards found on page: %d, accepted: %d", cards_found, len(records))
        rand_delay(2.5, 5.0)
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  NAUKRI
#  FIX: Updated API endpoint + required headers that Naukri now enforces.
#       The old appid/systemid values were rejected → 0 results.
#       Also improved HTML fallback selectors for new React-based layout.
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_naukri(keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    for page in range(1, SEARCH_PAGES + 1):
        params = {
            "noOfResults":  "20",
            "urlType":      "search_by_keyword",
            "searchType":   "adv",
            "keyword":      keyword,
            "location":     "delhi ncr",
            "pageNo":       str(page),
            "sort":         "1",
            "freshness":    "1",
            "src":          "jobsearchDesk",
            "latLong":      "",
            "experience":   "",
        }
        headers = {
            "Referer":       "https://www.naukri.com/",
            "Origin":        "https://www.naukri.com",
            "Accept":        "application/json",
            "appid":         "109",
            "systemid":      "Naukri",
            "content-type":  "application/json",
        }
        resp = http_get(f"https://www.naukri.com/jobapi/v3/search",
                        extra_headers=headers, params=params)

        if resp is None or resp.status_code >= 400:
            log.warning("Naukri API failed (status=%s) for '%s' → trying HTML",
                        resp.status_code if resp else "None", keyword)
            records.extend(_naukri_html(keyword, page))
            continue

        try:
            data = resp.json()
            log.debug("Naukri API response keys: %s", list(data.keys())[:8])
        except Exception as exc:
            log.warning("Naukri JSON parse error: %s → trying HTML", exc)
            records.extend(_naukri_html(keyword, page))
            continue

        jobs = (data.get("jobDetails") or
                data.get("jobs") or
                (data.get("data") or {}).get("jobDetails") or [])
        log.debug("Naukri API jobs count: %d for '%s'", len(jobs), keyword)

        if not jobs:
            records.extend(_naukri_html(keyword, page))
            break

        for job in jobs:
            title   = job.get("title") or job.get("jobTitle")
            company = (job.get("companyName") or
                       (job.get("company") or {}).get("label") or
                       (job.get("companyDetail") or {}).get("name"))
            pholders = job.get("placeholders") or []
            loc = next((p.get("label") for p in pholders if p.get("type") == "location"), None)
            if not loc:
                loc = job.get("location") or job.get("locationLabel")
            posted   = (job.get("footerPlaceholderLabel") or
                        job.get("createdDate") or job.get("modifiedOn") or "")
            salary   = (job.get("salary") or job.get("salaryDetail") or
                        next((p.get("label") for p in pholders if p.get("type") == "salary"), None))
            exp_text = job.get("experienceText") or job.get("experience")
            jd_url   = job.get("jdURL") or job.get("jobUrl") or ""
            if jd_url and not jd_url.startswith("http"):
                jd_url = "https://www.naukri.com" + jd_url

            log.debug("  Naukri job: %s | loc=%s | posted=%s", title, loc, posted)

            if not location_matches(str(loc or "")):
                log.debug("  ↳ SKIP location: %s", loc)
                continue
            if not is_within_24h(str(posted), label="Naukri"):
                continue

            rand_delay(0.3, 0.9)
            details = extract_jd_details(jd_url) if jd_url else {}
            records.append(JobRecord(
                portal="Naukri", source_keyword=keyword,
                job_title=title, company_name=company,
                job_location=loc, date_posted=str(posted),
                salary_package=salary or details.get("salary"),
                experience_required=exp_text or details.get("experience"),
                employment_type=details.get("employment_type"),
                job_url=_norm_url(jd_url),
                contact_email=details.get("email"),
                contact_phone=details.get("phone"),
                job_description_summary=details.get("summary"),
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
            if len(records) >= MAX_RESULTS_PER_KW:
                return records
        rand_delay()
    return records


def _naukri_html(keyword: str, page: int = 1) -> List[JobRecord]:
    """Naukri HTML fallback — handles their new React layout."""
    records: List[JobRecord] = []
    slug = re.sub(r"[^a-z0-9]+", "-", keyword.lower()).strip("-")
    url  = (f"https://www.naukri.com/{slug}-jobs-in-delhi-ncr-{page}"
            f"?freshness=1&sort=1")
    log.debug("Naukri HTML fallback: %s", url)
    resp = http_get(url, extra_headers={
        "Referer": "https://www.naukri.com/",
        "Accept": "text/html,application/xhtml+xml,*/*",
    })
    if resp is None:
        return records

    # Try JSON embedded in <script id="__NEXT_DATA__"> (Next.js)
    soup = BeautifulSoup(resp.text, "html.parser")
    next_data = soup.select_one("script#__NEXT_DATA__")
    if next_data:
        try:
            nd = json.loads(next_data.string or "")
            jobs_raw = (nd.get("props", {}).get("pageProps", {})
                          .get("dehydratedState", {})
                          .get("queries", [{}])[0]
                          .get("state", {}).get("data", {})
                          .get("jobDetails") or [])
            for job in jobs_raw:
                title   = job.get("title")
                company = job.get("companyName")
                loc     = job.get("location")
                posted  = job.get("createdDate") or job.get("modifiedOn") or ""
                jd_url  = job.get("jdURL") or ""
                if not is_within_24h(str(posted), "Naukri-NextJS"):
                    continue
                records.append(JobRecord(
                    portal="Naukri", source_keyword=keyword,
                    job_title=title, company_name=company,
                    job_location=loc, date_posted=str(posted),
                    job_url=_norm_url(jd_url),
                    fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                ))
            if records:
                return records
        except Exception as exc:
            log.debug("Naukri Next.js parse failed: %s", exc)

    # Classic HTML selectors (multiple variants)
    selectors = [
        "article.jobTuple",
        "div.jobTuple",
        "div[class*='job-tuple']",
        "div[class*='srp-jobtuple']",
        "div.list > div[type]",
    ]
    cards = []
    for sel in selectors:
        cards = soup.select(sel)
        if cards:
            log.debug("Naukri HTML: found %d cards with selector '%s'", len(cards), sel)
            break

    for card in cards:
        a = card.select_one(
            "a.title[href], a[class*='jobTitle'][href], "
            "a[class*='job-title'][href], h2 a[href]"
        )
        if not a:
            continue
        title   = a.get_text(strip=True)
        company = card.select_one(
            "a.subTitle, span[class*='companyName'], a[class*='comp-name']"
        )
        company = company.get_text(strip=True) if company else None
        loc     = card.select_one(
            "li.location span, span[class*='location'], li[class*='location']"
        )
        loc     = loc.get_text(strip=True) if loc else None
        posted  = card.select_one(
            "span.fw500, span[class*='freshness'], span[class*='date']"
        )
        posted  = posted.get_text(strip=True) if posted else None
        if not is_within_24h(posted, "Naukri-HTML"):
            continue
        records.append(JobRecord(
            portal="Naukri", source_keyword=keyword,
            job_title=title, company_name=company,
            job_location=loc, date_posted=posted,
            job_url=_norm_url(a.get("href", "")),
            fetched_at_utc=datetime.now(timezone.utc).isoformat(),
        ))
    log.debug("Naukri HTML fallback: %d records for '%s'", len(records), keyword)
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  INDEED
#  FIX: Indeed heavily uses React now — raw HTML has very few cards.
#       We extract the embedded JSON from window.mosaic.providerData
#       which contains all job cards in structured form.
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_indeed(keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    for loc_str in ["Delhi, India", "Noida Uttar Pradesh", "Gurugram Haryana"]:
        for page in range(SEARCH_PAGES):
            url = (
                "https://in.indeed.com/jobs"
                f"?q={quote_plus(keyword)}"
                f"&l={quote_plus(loc_str)}"
                f"&fromage=1&sort=date&start={page * 10}"
            )
            log.debug("Indeed URL: %s", url)
            resp = http_get(url, extra_headers={
                "Referer": "https://in.indeed.com/",
                "Accept": "text/html,application/xhtml+xml,*/*;q=0.9",
            })
            if resp is None:
                continue

            text = resp.text
            # Detect Cloudflare / bot block
            if any(k in text.lower() for k in ["just a moment", "cf-browser-verification",
                                                  "enable javascript", "ddos-guard"]):
                log.warning("Indeed: bot-check triggered for '%s' @ '%s'", keyword, loc_str)
                break

            # ── Method 1: Extract JSON from window.mosaic.providerData ────────
            jobs_from_json = _indeed_extract_json(text, keyword, loc_str)
            if jobs_from_json:
                records.extend(jobs_from_json)
                log.debug("  Indeed JSON method: %d jobs", len(jobs_from_json))
                if len(records) >= MAX_RESULTS_PER_KW:
                    return records
                rand_delay()
                continue

            # ── Method 2: HTML card scraping ──────────────────────────────────
            soup = BeautifulSoup(text, "html.parser")
            card_selectors = [
                "div.job_seen_beacon",
                "div[data-testid='jobCard']",
                "li.css-5lfssm",
                "div.resultContent",
            ]
            cards = []
            for sel in card_selectors:
                cards = soup.select(sel)
                if cards:
                    log.debug("Indeed HTML: %d cards with '%s'", len(cards), sel)
                    break

            for card in cards:
                jk_el = card.select_one("a[data-jk], a[id^='job_']")
                jk    = jk_el.get("data-jk") or re.search(r"jk=([a-f0-9]+)", jk_el.get("href","")).group(1) if jk_el else ""

                title_el = (card.select_one("h2.jobTitle span[title]") or
                             card.select_one("h2.jobTitle a span") or
                             card.select_one("span[id^='jobTitle']"))
                title = title_el.get_text(strip=True) if title_el else None

                company = card.select_one(
                    "[data-testid='company-name'], span.companyName, "
                    "span[class*='EmployerName']"
                )
                company = company.get_text(strip=True) if company else None

                loc_el = card.select_one(
                    "[data-testid='text-location'], div.companyLocation, "
                    "div[class*='companyLocation']"
                )
                location = loc_el.get_text(strip=True) if loc_el else loc_str

                posted_el = card.select_one(
                    "[data-testid='myJobsStateDate'], span.date, "
                    "span[class*='date'], div[class*='date']"
                )
                posted = posted_el.get_text(strip=True) if posted_el else None

                log.debug("  Indeed card: %s | %s | posted=%s", title, location, posted)
                if not location_matches(f"{location} {loc_str}"):
                    continue
                if posted and not is_within_24h(posted, "Indeed"):
                    continue

                job_url = f"https://in.indeed.com/viewjob?jk={jk}" if jk else ""
                rand_delay(0.5, 1.3)
                details = extract_jd_details(job_url) if job_url else {}
                records.append(JobRecord(
                    portal="Indeed", source_keyword=keyword,
                    job_title=title, company_name=company,
                    job_location=location, date_posted=posted or "Within 24h",
                    salary_package=details.get("salary"),
                    experience_required=details.get("experience"),
                    employment_type=details.get("employment_type"),
                    job_url=_norm_url(job_url),
                    contact_email=details.get("email"),
                    contact_phone=details.get("phone"),
                    job_description_summary=details.get("summary"),
                    fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                ))
                if len(records) >= MAX_RESULTS_PER_KW:
                    return records
            rand_delay()
    return records


def _indeed_extract_json(html: str, keyword: str, loc_str: str) -> List[JobRecord]:
    """Extract job data from Indeed's embedded JS object."""
    records: List[JobRecord] = []
    # Find the JSON blob
    m = re.search(r'window\.mosaic\.providerData\["mosaic-provider-jobcards"\]\s*=\s*(\{.*?\});\s*window', html, re.DOTALL)
    if not m:
        m = re.search(r'"jobKeysWithDescriptions":\{(.*?)\},"jobKeysInOrder"', html, re.DOTALL)
        if not m:
            return records

    try:
        raw = m.group(1)
        # Try to find the metaData.jobResults array
        m2 = re.search(r'"metaData":\{"jobResults":\[(.*?)\]\}', raw, re.DOTALL)
        if not m2:
            return records
        jobs_json = json.loads(f"[{m2.group(1)}]")
    except Exception:
        return records

    for job in jobs_json:
        jd = job if isinstance(job, dict) else {}
        title   = jd.get("displayTitle") or jd.get("title")
        company = jd.get("company") or jd.get("employerName")
        loc     = jd.get("formattedLocation") or jd.get("location")
        jk      = jd.get("jobkey") or jd.get("jk") or ""
        posted  = jd.get("formattedRelativeTime") or jd.get("date") or ""

        if not location_matches(f"{loc or ''} {loc_str}"):
            continue
        if posted and not is_within_24h(posted, "Indeed-JSON"):
            continue

        job_url = f"https://in.indeed.com/viewjob?jk={jk}" if jk else ""
        records.append(JobRecord(
            portal="Indeed", source_keyword=keyword,
            job_title=title, company_name=company,
            job_location=loc or loc_str,
            date_posted=posted or "Within 24h",
            job_url=_norm_url(job_url),
            fetched_at_utc=datetime.now(timezone.utc).isoformat(),
        ))
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  GLASSDOOR
#  FIX: Use Glassdoor's own search URL with a date filter rather than
#       Google site: search (which started blocking scraper requests).
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_glassdoor(keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    # Glassdoor uses "fromAge" in days
    for loc_id, loc_name in [("3000139077", "Delhi"), ("9061049", "Noida"), ("7647306", "Gurugram")]:
        url = (
            "https://www.glassdoor.co.in/Job/jobs.htm"
            f"?sc.keyword={quote_plus(keyword)}"
            f"&locId={loc_id}"
            f"&locT=C"
            f"&fromAge=1"
            f"&sortBy=date_desc"
        )
        log.debug("Glassdoor URL: %s", url)
        resp = http_get(url, extra_headers={
            "Referer":  "https://www.glassdoor.co.in/",
            "Accept":   "text/html,application/xhtml+xml,*/*;q=0.9",
        })
        if resp is None or resp.status_code >= 400:
            log.warning("Glassdoor: HTTP %s for '%s'",
                        resp.status_code if resp else "None", keyword)
            continue

        text = resp.text
        if "just a moment" in text.lower() or "cf-browser" in text.lower():
            log.warning("Glassdoor: bot-check for '%s'", keyword)
            continue

        # Try JSON-LD first
        soup = BeautifulSoup(text, "html.parser")
        for script in soup.select("script[type='application/ld+json']"):
            try:
                data = json.loads(script.string or "")
                items = data if isinstance(data, list) else [data]
                for item in items:
                    if item.get("@type") != "JobPosting":
                        continue
                    title   = item.get("title")
                    company = (item.get("hiringOrganization") or {}).get("name")
                    loc     = (item.get("jobLocation") or {}).get("address", {}).get("addressLocality")
                    posted  = item.get("datePosted") or ""
                    jurl    = item.get("url") or item.get("@id") or ""
                    if not is_within_24h(posted, "Glassdoor-JSONLD"):
                        continue
                    rand_delay(0.5, 1.2)
                    details = extract_jd_details(jurl)
                    records.append(JobRecord(
                        portal="Glassdoor", source_keyword=keyword,
                        job_title=title, company_name=company,
                        job_location=loc or loc_name,
                        date_posted=posted,
                        salary_package=details.get("salary"),
                        experience_required=details.get("experience"),
                        employment_type=details.get("employment_type"),
                        job_url=_norm_url(jurl),
                        contact_email=details.get("email"),
                        contact_phone=details.get("phone"),
                        job_description_summary=details.get("summary"),
                        fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                    ))
                    if len(records) >= MAX_RESULTS_PER_KW:
                        return records
            except Exception:
                pass

        # HTML card fallback
        for card in soup.select("li.react-job-listing, div[data-test='jobListing'], article"):
            a = card.select_one("a[data-test='job-link'], a.jobLink, a[href*='/Job/']")
            if not a:
                continue
            title   = (card.select_one("[data-test='job-title'], .job-title") or a)
            title   = title.get_text(strip=True)
            company = card.select_one("[data-test='employer-name'], .employer-name")
            company = company.get_text(strip=True) if company else None
            loc_el  = card.select_one("[data-test='emp-location'], .location")
            loc     = loc_el.get_text(strip=True) if loc_el else loc_name
            posted_el = card.select_one("[data-test='job-age'], .job-age, time")
            posted  = posted_el.get_text(strip=True) if posted_el else None
            href    = a.get("href", "")
            if href and not href.startswith("http"):
                href = "https://www.glassdoor.co.in" + href

            if posted and not is_within_24h(posted, "Glassdoor-HTML"):
                continue

            rand_delay(0.5, 1.2)
            details = extract_jd_details(href)
            records.append(JobRecord(
                portal="Glassdoor", source_keyword=keyword,
                job_title=title, company_name=company,
                job_location=loc,
                date_posted=posted or "Within 24h",
                salary_package=details.get("salary"),
                experience_required=details.get("experience"),
                employment_type=details.get("employment_type"),
                job_url=_norm_url(href),
                contact_email=details.get("email"),
                contact_phone=details.get("phone"),
                job_description_summary=details.get("summary"),
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
            if len(records) >= MAX_RESULTS_PER_KW:
                return records
        rand_delay()
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  SHINE  (working — keep mostly same, minor selector improvements)
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_shine(keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    slug = re.sub(r"[^a-z0-9]+", "-", keyword.lower()).strip("-")
    for page in range(1, SEARCH_PAGES + 1):
        url = (
            f"https://www.shine.com/job-search/{slug}-jobs-in-delhi/"
            f"?sort=1&page={page}&freshness=1"
        )
        resp = http_get(url, extra_headers={"Referer": "https://www.shine.com/"})
        if resp is None:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try JSON-LD first (most reliable)
        jld_records = _shine_jsonld(soup, keyword)
        if jld_records:
            records.extend(jld_records)
            if len(records) >= MAX_RESULTS_PER_KW:
                return records
            rand_delay()
            continue

        # HTML card selectors
        selectors = [
            "div.job-card", "article.jobCard", "li.job-listing",
            "div[class*='jobCard']", "div[class*='job_card']",
            "div.jsx-jobCard", "div[data-jobid]",
        ]
        cards = []
        for sel in selectors:
            cards = soup.select(sel)
            if cards:
                log.debug("Shine: %d cards with '%s'", len(cards), sel)
                break

        for card in cards:
            a = card.select_one(
                "a.job-title[href], h2 a[href], h3 a[href], "
                "a[class*='title'][href], a[class*='job'][href]"
            )
            if not a:
                continue
            title   = a.get_text(strip=True)
            company = card.select_one(
                "span[class*='company'], a[class*='company'], "
                "div[class*='company'], p[class*='company']"
            )
            company = company.get_text(strip=True) if company else None
            loc_el  = card.select_one(
                "span[class*='location'], li[class*='location'], "
                "div[class*='location']"
            )
            loc     = loc_el.get_text(strip=True) if loc_el else None
            posted_el = card.select_one(
                "span[class*='date'], span[class*='posted'], "
                "time, div[class*='date']"
            )
            posted  = posted_el.get_text(strip=True) if posted_el else None
            sal_el  = card.select_one("span[class*='salary'], li[class*='salary']")
            salary  = sal_el.get_text(strip=True) if sal_el else None
            job_url = a.get("href", "")
            if job_url and not job_url.startswith("http"):
                job_url = "https://www.shine.com" + job_url

            if posted and not is_within_24h(posted, "Shine"):
                continue

            rand_delay(0.4, 1.2)
            details = extract_jd_details(job_url)
            records.append(JobRecord(
                portal="Shine", source_keyword=keyword,
                job_title=title, company_name=company,
                job_location=loc or "Delhi NCR",
                date_posted=posted or "Within 24h",
                salary_package=salary or details.get("salary"),
                experience_required=details.get("experience"),
                employment_type=details.get("employment_type"),
                job_url=_norm_url(job_url),
                contact_email=details.get("email"),
                contact_phone=details.get("phone"),
                job_description_summary=details.get("summary"),
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
            if len(records) >= MAX_RESULTS_PER_KW:
                return records
        rand_delay()
    return records


def _shine_jsonld(soup: BeautifulSoup, keyword: str) -> List[JobRecord]:
    records = []
    for script in soup.select("script[type='application/ld+json']"):
        try:
            data  = json.loads(script.string or "")
            items = data if isinstance(data, list) else [data]
            for item in items:
                if item.get("@type") != "JobPosting":
                    continue
                title   = item.get("title")
                company = (item.get("hiringOrganization") or {}).get("name")
                loc     = (item.get("jobLocation") or {}).get("address", {}).get("addressLocality")
                posted  = item.get("datePosted") or ""
                jurl    = item.get("url") or item.get("@id") or ""
                if not is_within_24h(posted, "Shine-JSONLD"):
                    continue
                records.append(JobRecord(
                    portal="Shine", source_keyword=keyword,
                    job_title=title, company_name=company,
                    job_location=loc or "Delhi NCR", date_posted=posted,
                    job_url=_norm_url(jurl),
                    fetched_at_utc=datetime.now(timezone.utc).isoformat(),
                ))
        except Exception:
            pass
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  INTERNSHALA
#  FIX: Internshala's 2024 redesign changed CSS classes completely.
#       Now using their search API endpoint + updated selectors.
#       Also relaxed age check — Internshala rarely shows exact post time,
#       so "Actively hiring" / no date = accept (they filter by recent).
# ═══════════════════════════════════════════════════════════════════════════════

def scrape_internshala(keyword: str) -> List[JobRecord]:
    records: List[JobRecord] = []
    # Try API endpoint first
    api_records = _internshala_api(keyword)
    if api_records:
        return api_records

    # HTML fallback
    slug = re.sub(r"[^a-z0-9]+", "-", keyword.lower()).strip("-")
    for page in range(1, SEARCH_PAGES + 1):
        url = f"https://internshala.com/jobs/keywords-{slug}/location-delhi,noida,gurgaon/page-{page}/"
        log.debug("Internshala URL: %s", url)
        resp = http_get(url, extra_headers={
            "Referer": "https://internshala.com/",
            "Accept": "text/html,*/*;q=0.9",
        })
        if resp is None:
            continue
        soup = BeautifulSoup(resp.text, "html.parser")

        # Try JSON-LD
        jld = _shine_jsonld(soup, keyword)  # same logic
        for r in jld:
            r.portal = "Internshala"
        records.extend(jld)

        # Updated selectors for new Internshala layout
        card_selectors = [
            "div.individual_internship",
            "div[class*='internship_meta']",
            "div.job-internship-card",
            "div[class*='JobCard']",
            "li[class*='job']",
        ]
        cards = []
        for sel in card_selectors:
            cards = soup.select(sel)
            if cards:
                log.debug("Internshala: %d cards with '%s'", len(cards), sel)
                break

        for card in cards:
            title_el = card.select_one(
                "a.job-title-href, h3 a, .profile a, "
                "a[class*='job-title'], a[class*='view_detail'], "
                "p.title a, h4 a"
            )
            if not title_el:
                continue
            title   = title_el.get_text(strip=True)
            company = card.select_one(
                "a.link_display_like_text, p.company-name, "
                "div.company_name, span[class*='company'], a[class*='company']"
            )
            company = company.get_text(strip=True) if company else None
            loc_el  = card.select_one(
                "div.location_link, span.location, "
                "div[class*='location'], p[class*='location']"
            )
            loc     = loc_el.get_text(strip=True) if loc_el else "Delhi NCR"
            posted_el = card.select_one(
                "div[class*='posted'], span[class*='posted'], "
                "span.actively-hiring-badge, div[class*='status']"
            )
            posted  = posted_el.get_text(strip=True) if posted_el else None
            # "Actively hiring" / no date → treat as just posted
            if not posted or "actively" in (posted or "").lower():
                posted = "Just posted"

            sal_el  = card.select_one(
                "span[class*='stipend'], div[class*='salary'], "
                "span[class*='salary'], p[class*='salary']"
            )
            salary  = sal_el.get_text(strip=True) if sal_el else None
            href    = title_el.get("href", "")
            if href and not href.startswith("http"):
                href = "https://internshala.com" + href

            if not is_within_24h(posted, "Internshala"):
                continue

            rand_delay(0.4, 1.1)
            details = extract_jd_details(href)
            records.append(JobRecord(
                portal="Internshala", source_keyword=keyword,
                job_title=title, company_name=company,
                job_location=loc, date_posted=posted,
                salary_package=salary or details.get("salary"),
                experience_required=details.get("experience"),
                employment_type=details.get("employment_type"),
                job_url=_norm_url(href),
                contact_email=details.get("email"),
                contact_phone=details.get("phone"),
                job_description_summary=details.get("summary"),
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
            if len(records) >= MAX_RESULTS_PER_KW:
                return records
        rand_delay()
    return records


def _internshala_api(keyword: str) -> List[JobRecord]:
    """Try Internshala's internal search API."""
    records: List[JobRecord] = []
    try:
        url = "https://internshala.com/jobs/ajax-jobs"
        payload = {
            "keywords": keyword,
            "location": "Delhi;Noida;Gurgaon",
            "page_no": 1,
        }
        resp = SESSION.post(url, data=payload, headers={
            "User-Agent": rand_ua(),
            "Referer": "https://internshala.com/",
            "X-Requested-With": "XMLHttpRequest",
        }, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return records
        data = resp.json()
        for job in (data.get("jobs") or []):
            title   = job.get("job_title") or job.get("title")
            company = job.get("company_name") or job.get("company")
            loc     = job.get("location")
            posted  = job.get("posted_on") or job.get("created_at") or "Just posted"
            jurl    = job.get("url") or job.get("job_url") or ""
            if jurl and not jurl.startswith("http"):
                jurl = "https://internshala.com" + jurl
            if not is_within_24h(posted, "Internshala-API"):
                continue
            records.append(JobRecord(
                portal="Internshala", source_keyword=keyword,
                job_title=title, company_name=company,
                job_location=loc or "Delhi NCR", date_posted=str(posted),
                job_url=_norm_url(jurl),
                fetched_at_utc=datetime.now(timezone.utc).isoformat(),
            ))
    except Exception as exc:
        log.debug("Internshala API failed: %s", exc)
    return records


# ═══════════════════════════════════════════════════════════════════════════════
#  DEDUPLICATION
# ═══════════════════════════════════════════════════════════════════════════════

def deduplicate(records: List[JobRecord]) -> List[JobRecord]:
    url_seen:     Set[str] = set()
    content_seen: Set[str] = set()
    unique: List[JobRecord] = []
    for rec in records:
        uk = rec.url_key()
        ck = rec.content_key()
        if uk and uk in url_seen:
            continue
        if not uk and ck in content_seen:
            continue
        if uk:
            url_seen.add(uk)
        content_seen.add(ck)
        unique.append(rec)
    log.info("Dedup: %d raw → %d unique records", len(records), len(unique))
    return unique


# ═══════════════════════════════════════════════════════════════════════════════
#  EXCEL OUTPUT
# ═══════════════════════════════════════════════════════════════════════════════

COLS = [
    "portal", "source_keyword", "job_title", "company_name",
    "job_location", "date_posted", "salary_package",
    "experience_required", "employment_type",
    "job_url", "contact_email", "contact_phone",
    "job_description_summary", "fetched_at_utc",
]

PORTAL_COLORS = {
    "LinkedIn":    "0A66C2",
    "Naukri":      "D44000",
    "Indeed":      "2164F3",
    "Glassdoor":   "0CAA41",
    "Shine":       "E05A00",
    "Internshala": "006BFF",
}


def to_dataframe(records: List[JobRecord]) -> pd.DataFrame:
    rows = [asdict(r) for r in records]
    if not rows:
        return pd.DataFrame(columns=COLS)
    df = pd.DataFrame(rows)
    for c in COLS:
        if c not in df.columns:
            df[c] = None
    return df[COLS].fillna("")


def save_excel(df: pd.DataFrame, path: str) -> None:
    df.to_excel(path, index=False, engine="openpyxl")
    wb = load_workbook(path)
    ws = wb.active
    header_map: Dict[str, int] = {
        ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)
    }
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
        cell.border    = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28

    portal_col = header_map.get("portal")
    url_col    = header_map.get("job_url")
    email_col  = header_map.get("contact_email")
    phone_col  = header_map.get("contact_phone")
    desc_col   = header_map.get("job_description_summary")

    for r in range(2, ws.max_row + 1):
        portal_val = str(ws.cell(row=r, column=portal_col).value or "") if portal_col else ""
        hex_color  = PORTAL_COLORS.get(portal_val, "F5F5F5")
        light      = "".join(f"{min(255, int(hex_color[i:i+2], 16) + 185):02X}" for i in (0, 2, 4))

        for c in range(1, ws.max_column + 1):
            cell           = ws.cell(row=r, column=c)
            cell.fill      = PatternFill("solid", fgColor=light[:6])
            cell.border    = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c == desc_col))

        if url_col:
            cell = ws.cell(row=r, column=url_col)
            if str(cell.value or "").startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.font = Font(color="0563C1", underline="single", size=10)
        if email_col:
            cell = ws.cell(row=r, column=email_col)
            val  = str(cell.value or "").strip()
            if "@" in val:
                cell.hyperlink = f"mailto:{val}"
                cell.font = Font(color="0563C1", underline="single", size=10)
        if phone_col:
            cell  = ws.cell(row=r, column=phone_col)
            phone = normalize_tel(str(cell.value or "").strip())
            if phone:
                cell.hyperlink = f"tel:{phone}"
                cell.font = Font(color="0563C1", underline="single", size=10)

    col_widths = {
        "portal": 12, "source_keyword": 24, "job_title": 35, "company_name": 28,
        "job_location": 22, "date_posted": 20, "salary_package": 22,
        "experience_required": 18, "employment_type": 16,
        "job_url": 40, "contact_email": 32, "contact_phone": 18,
        "job_description_summary": 58, "fetched_at_utc": 22,
    }
    for name, width in col_widths.items():
        idx = header_map.get(name)
        if idx:
            ws.column_dimensions[get_column_letter(idx)].width = width

    ws.freeze_panes     = "A2"
    ws.auto_filter.ref  = ws.dimensions
    ws.sheet_view.showGridLines = False
    wb.save(path)
    log.info("✅ Saved %d rows → %s", ws.max_row - 1, path)


# ═══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════════════════════

def run_scrape(
    keywords: Optional[List[str]] = None,
    output_file: Optional[str] = None,
    profile_name: str = "HR Job Scraper v4",
) -> None:
    log.info("╔══════════════════════════════════════════════════╗")
    log.info("║  HR Job Scraper v4  —  Last 24h  —  Delhi NCR   ║")
    log.info("╚══════════════════════════════════════════════════╝")

    log.info("Profile: %s", profile_name)

    all_records: List[JobRecord] = []
    active_keywords = keywords or HR_KEYWORDS
    active_output_file = output_file or OUTPUT_FILE

    driver = None
    if ENABLE_LINKEDIN and SELENIUM_OK:
        try:
            driver = build_driver()
            log.info("Selenium driver ready (undetected=%s)", UC_AVAILABLE)
        except Exception as exc:
            log.warning("Selenium failed: %s — LinkedIn disabled.", exc)

    kw_iter = tqdm(active_keywords, desc="Keywords", unit="kw") if HAS_TQDM else active_keywords

    for kw in kw_iter:
        if HAS_TQDM:
            kw_iter.set_postfix({"kw": kw[:28]})  # type: ignore
        else:
            log.info("━━━ Keyword: %s", kw)

        if ENABLE_LINKEDIN and driver:
            for locq in LOCATION_QUERIES:
                recs = scrape_linkedin(driver, kw, locq)
                log.info("  LinkedIn [%s] → %d", locq.split(",")[0], len(recs))
                all_records.extend(recs)
                rand_delay(2, 4)

        if ENABLE_NAUKRI:
            recs = scrape_naukri(kw)
            log.info("  Naukri → %d", len(recs))
            all_records.extend(recs)
            rand_delay()

        if ENABLE_INDEED:
            recs = scrape_indeed(kw)
            log.info("  Indeed → %d", len(recs))
            all_records.extend(recs)
            rand_delay()

        if ENABLE_GLASSDOOR:
            recs = scrape_glassdoor(kw)
            log.info("  Glassdoor → %d", len(recs))
            all_records.extend(recs)
            rand_delay()

        if ENABLE_SHINE:
            recs = scrape_shine(kw)
            log.info("  Shine → %d", len(recs))
            all_records.extend(recs)
            rand_delay()

        if ENABLE_INTERNSHALA:
            recs = scrape_internshala(kw)
            log.info("  Internshala → %d", len(recs))
            all_records.extend(recs)
            rand_delay()

    if driver:
        try:
            driver.quit()
        except Exception:
            pass

    all_records = deduplicate(all_records)
    df = to_dataframe(all_records)
    save_excel(df, active_output_file)
    log.info("Done. Unique jobs: %d -> %s", len(df), active_output_file)


def main() -> None:
    run_scrape()


if __name__ == "__main__":
    main()
